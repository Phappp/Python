from app.models.user_model import User
from app.models.course_model import Course
from flask import request, redirect, url_for, flash, render_template, send_file
from werkzeug.security import generate_password_hash
from datetime import datetime
import io
import openpyxl

class AdminController:
    @staticmethod
    def list_users():
        users = list(User.get_all())
        return render_template('admin/users.html', users=users)

    @staticmethod
    def add_user(form):
        if request.method == 'POST' and form.validate_on_submit():
            User.create(
                username=form.username.data,
                email=form.email.data,
                password=form.password.data,
                role=form.role.data
            )
            flash('Tạo tài khoản thành công!', 'success')
            return redirect(url_for('admin.list_users'))
        return render_template('admin/add_user.html', form=form)

    @staticmethod
    def edit_user(username, form):
        user = User.find_by_username(username)
        if not user:
            flash('Không tìm thấy người dùng!', 'danger')
            return redirect(url_for('admin.list_users'))
        if request.method == 'POST' and form.validate_on_submit():
            update_data = {
                'email': form.email.data,
                'full_name': form.full_name.data,
                'role': form.role.data,
                'updated_at': datetime.utcnow()
            }
            User.update_profile(username, update_data)
            flash('Cập nhật thông tin thành công!', 'success')
            return redirect(url_for('admin.list_users'))
        return render_template('admin/edit_user.html', form=form, user=user)

    @staticmethod
    def delete_user(username):
        User.delete(username)
        flash('Xóa tài khoản thành công!', 'success')
        return redirect(url_for('admin.list_users'))

    @staticmethod
    def change_role(username, new_role):
        User.update_profile(username, {'role': new_role, 'updated_at': datetime.utcnow()})
        flash('Cập nhật vai trò thành công!', 'success')
        return redirect(url_for('admin.list_users'))

    @staticmethod
    def reset_password(username, new_password):
        User.update_password(username, new_password)
        flash('Đặt lại mật khẩu thành công!', 'success')
        return redirect(url_for('admin.list_users'))

    @staticmethod
    def toggle_active(username, active):
        User.update_profile(username, {'active': active, 'updated_at': datetime.utcnow()})
        flash(('Mở khóa' if active else 'Khóa') + ' tài khoản thành công!', 'success')
        return redirect(url_for('admin.list_users'))

    @staticmethod
    def list_courses():
        courses = list(Course.get_all())
        return render_template('admin/courses.html', courses=courses)

    @staticmethod
    def add_course(form):
        if request.method == 'POST' and form.validate_on_submit():
            Course.create(
                name=form.name.data,
                description=form.description.data,
                original_price=form.original_price.data,
                discounted_price=form.discounted_price.data,
                created_by='admin',  # Ghi nhận admin tạo
                color=form.color.data,
                icon_class=form.icon_class.data
            )
            flash('Tạo khóa học thành công!', 'success')
            return redirect(url_for('admin.list_courses'))
        return render_template('admin/add_course.html', form=form)

    @staticmethod
    def edit_course(course_id, form):
        course = Course.find_by_id(course_id)
        if not course:
            flash('Không tìm thấy khóa học!', 'danger')
            return redirect(url_for('admin.list_courses'))
        if request.method == 'POST' and form.validate_on_submit():
            update_data = {
                'name': form.name.data,
                'description': form.description.data,
                'original_price': form.original_price.data,
                'discounted_price': form.discounted_price.data,
                'color': form.color.data,
                'icon_class': form.icon_class.data
            }
            Course.update(course_id, update_data)
            flash('Cập nhật khóa học thành công!', 'success')
            return redirect(url_for('admin.list_courses'))
        return render_template('admin/edit_course.html', form=form, course=course)

    @staticmethod
    def delete_course(course_id):
        Course.delete(course_id)
        flash('Xóa khóa học thành công!', 'success')
        return redirect(url_for('admin.list_courses'))

    @staticmethod
    def view_course_students(course_id):
        course = Course.find_by_id(course_id)
        students = course.get('students', []) if course else []
        return render_template('admin/course_students.html', course=course, students=students)

    @staticmethod
    def stats():
        from app.models.user_model import User
        from app.models.course_model import Course
        users = list(User.get_all())
        courses = list(Course.get_all())
        num_students = sum(1 for u in users if u.get('role') == 'student')
        num_lectures = sum(1 for u in users if u.get('role') == 'lecture')
        num_admins = sum(1 for u in users if u.get('role') == 'admin')
        num_courses = len(courses)
        # Tỷ lệ hoàn thành bài tập: cần có logic submissions, tạm để None
        completion_rate = None
        return render_template('admin/stats.html', num_students=num_students, num_lectures=num_lectures, num_admins=num_admins, num_courses=num_courses, completion_rate=completion_rate)

    @staticmethod
    def export_excel():
        from app.models.user_model import User
        from app.models.course_model import Course
        users = list(User.get_all())
        courses = list(Course.get_all())
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Thống kê chung'
        ws1.append(['Loại', 'Số lượng'])
        num_students = sum(1 for u in users if u.get('role') == 'student')
        num_lectures = sum(1 for u in users if u.get('role') == 'lecture')
        num_admins = sum(1 for u in users if u.get('role') == 'admin')
        num_courses = len(courses)
        ws1.append(['Học sinh', num_students])
        ws1.append(['Giáo viên', num_lectures])
        ws1.append(['Admin', num_admins])
        ws1.append(['Khóa học', num_courses])
        # Sheet user
        ws2 = wb.create_sheet('Danh sách user')
        ws2.append(['Username', 'Email', 'Role', 'Trạng thái'])
        for u in users:
            ws2.append([
                u.get('username'),
                u.get('email'),
                u.get('role'),
                'Hoạt động' if u.get('active', True) else 'Tạm khóa'
            ])
        # Sheet course
        ws3 = wb.create_sheet('Danh sách khóa học')
        ws3.append(['Tên khóa học', 'Mô tả', 'Người tạo', 'Số học sinh đăng ký'])
        for c in courses:
            ws3.append([
                c.get('name'),
                c.get('description'),
                c.get('created_by'),
                len(c.get('students', []))
            ])
        # Xuất file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='bao_cao_thong_ke.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @staticmethod
    def roles():
        # Quyền mặc định cho từng role
        roles_permissions = {
            'admin': ['Quản lý người dùng', 'Quản lý khóa học', 'Thống kê & báo cáo', 'Cấu hình hệ thống'],
            'lecture': ['Quản lý khóa học', 'Quản lý bài tập'],
            'student': ['Xem khóa học', 'Làm bài tập']
        }
        return render_template('admin/roles.html', roles_permissions=roles_permissions)

    @staticmethod
    def admin_dashboard():
        from app.models.user_model import User
        from app.models.course_model import Course
        from app.models.exercise_model import Exercise
        users = list(User.get_all())
        courses = list(Course.get_all())
        try:
            from app.models.exercise_model import Exercise
            exercises = list(Exercise.get_all())
        except Exception:
            exercises = []
        num_students = sum(1 for u in users if u.get('role') == 'student')
        num_lectures = sum(1 for u in users if u.get('role') == 'lecture')
        num_admins = sum(1 for u in users if u.get('role') == 'admin')
        num_locked = sum(1 for u in users if not u.get('active', True))
        num_courses = len(courses)
        num_exercises = len(exercises)
        return render_template('admin/dashboard.html',
            num_students=num_students,
            num_lectures=num_lectures,
            num_admins=num_admins,
            num_locked=num_locked,
            num_courses=num_courses,
            num_exercises=num_exercises
        )

class PermissionUserController:
    @staticmethod
    def users_by_role(role_name):
        users = [u for u in User.get_all() if u.get('role') == role_name]
        return render_template('admin/users_by_role.html', users=users, role_name=role_name)

    @staticmethod
    def view_user_permissions(username):
        user = User.find_by_username(username)
        if not user:
            flash('Không tìm thấy người dùng!', 'danger')
            return redirect(url_for('admin.list_users'))
        # Quyền mặc định từ role
        role = user.get('role')
        # Quyền riêng
        user_permissions = user.get('permissions', [])
        # Gợi ý một số quyền phổ biến
        all_permissions = ['manage_users', 'manage_courses', 'manage_exercises', 'view_stats', 'config_system']
        return render_template('admin/user_permissions.html', user=user, role=role, user_permissions=user_permissions, all_permissions=all_permissions)

    @staticmethod
    def update_user_permissions(username):
        user = User.find_by_username(username)
        if not user:
            flash('Không tìm thấy người dùng!', 'danger')
            return redirect(url_for('admin.list_users'))
        perms = request.form.getlist('permissions')
        User.set_permissions(username, perms)
        flash('Cập nhật quyền cho tài khoản thành công!', 'success')
        return redirect(url_for('admin.view_user_permissions', username=username)) 