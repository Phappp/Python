from flask import session, current_app
from app.models.user_model import User
from app.models.course_model import Course
from app.models.exercise_model import Exercise, Submission
from app.models.quiz_model import Quiz, QuizSubmission
from app.models.document_model import Document
from datetime import datetime, timedelta
import json
from bson import ObjectId

class StatisticsController:
    
    @staticmethod
    def get_lecture_dashboard_stats():
        """
        Lấy thống kê tổng quan cho lecture dashboard
        """
        username = session.get('username')
        role = session.get('role')
        if not username:
            return None
            
        try:
            # Thống kê khóa học
            if role == 'lecture':
                courses = list(Course.get_all())
            else:
                courses = list(Course.find_by_creator(username))
            total_courses = len(courses)
            active_courses = len([c for c in courses if c.get('is_active', True)])
            
            # Thống kê sinh viên
            if role == 'lecture':
                total_students = User.count_by_role('student')
            else:
                total_students = 0
                for course in courses:
                    total_students += len(course.get('students', []))
            
            # Thống kê bài tập
            if role == 'lecture':
                exercises = list(Exercise.get_all())
            else:
                exercises = list(Exercise.find_by_creator(username))
            total_exercises = len(exercises)
            active_exercises = len([e for e in exercises if e.get('is_visible', True)])
            
            # Thống kê số lượt nộp bài
            total_submissions = 0
            for exercise in exercises:
                submissions = list(Submission.find_by_exercise(str(exercise['_id'])))
                total_submissions += len(submissions)
            
            # Thống kê quiz
            if role == 'lecture':
                quizzes = list(Quiz.get_all())
            else:
                quizzes = list(Quiz.find_by_creator(username))
            total_quizzes = len(quizzes)
            total_quiz_submissions = 0
            for quiz in quizzes:
                quiz_submissions = list(QuizSubmission.find_by_quiz(str(quiz['_id'])))
                total_quiz_submissions += len(quiz_submissions)
            
            # Thống kê tài liệu
            if role == 'lecture':
                documents = list(Document.get_all())
            else:
                documents = list(Document.find_by_uploader(username))
            total_documents = len(documents)
            total_downloads = sum(d.get('download_count', 0) for d in documents)
            
            # Thống kê điểm số (trung bình)
            avg_score = 0
            total_scores = 0
            score_count = 0
            
            for exercise in exercises:
                submissions = list(Submission.find_by_exercise(str(exercise['_id'])))
                for submission in submissions:
                    if submission.get('score') is not None:
                        total_scores += submission['score']
                        score_count += 1
            
            if score_count > 0:
                avg_score = round(total_scores / score_count, 1)
            
            # Thống kê theo thời gian (7 ngày gần nhất)
            last_7_days = datetime.utcnow() - timedelta(days=7)
            recent_submissions = 0
            recent_quiz_submissions = 0
            
            for exercise in exercises:
                submissions = list(Submission.find_by_exercise(str(exercise['_id'])))
                for submission in submissions:
                    if submission.get('submitted_at') and submission['submitted_at'] >= last_7_days:
                        recent_submissions += 1
            
            for quiz in quizzes:
                quiz_submissions = list(QuizSubmission.find_by_quiz(str(quiz['_id'])))
                for submission in quiz_submissions:
                    if submission.get('submitted_at') and submission['submitted_at'] >= last_7_days:
                        recent_quiz_submissions += 1
            
            return {
                'courses': {
                    'total': total_courses,
                    'active': active_courses
                },
                'students': {
                    'total': total_students
                },
                'exercises': {
                    'total': total_exercises,
                    'active': active_exercises
                },
                'submissions': {
                    'total': total_submissions,
                    'recent_7_days': recent_submissions
                },
                'quizzes': {
                    'total': total_quizzes,
                    'submissions': total_quiz_submissions,
                    'recent_7_days': recent_quiz_submissions
                },
                'documents': {
                    'total': total_documents,
                    'downloads': total_downloads
                },
                'performance': {
                    'avg_score': avg_score,
                    'total_scores': score_count
                }
            }
        except Exception as e:
            current_app.logger.error(f"Error getting lecture stats: {e}")
            return None
    
    @staticmethod
    def get_course_progress_stats(course_id):
        """
        Lấy thống kê tiến độ cho một khóa học cụ thể
        """
        try:
            course = Course.find_by_id(course_id)
            if not course:
                return None
                
            students = course.get('students', [])
            exercises = list(Exercise.find_by_course(course_id))
            
            # Thống kê bài tập theo khóa học
            exercise_stats = []
            for exercise in exercises:
                submissions = list(Submission.find_by_exercise(str(exercise['_id'])))
                total_submissions = len(submissions)
                completed_submissions = len([s for s in submissions if s.get('score') is not None])
                avg_score = 0
                
                if completed_submissions > 0:
                    total_score = sum(s['score'] for s in submissions if s.get('score') is not None)
                    avg_score = round(total_score / completed_submissions, 1)
                
                exercise_stats.append({
                    'exercise_id': str(exercise['_id']),
                    'title': exercise.get('title'),
                    'total_submissions': total_submissions,
                    'completed_submissions': completed_submissions,
                    'completion_rate': round((completed_submissions / len(students)) * 100, 1) if students else 0,
                    'avg_score': avg_score
                })
            
            # Thống kê sinh viên
            student_stats = []
            for student_username in students:
                student = User.find_by_username(student_username)
                if student:
                    completed_exercises = 0
                    total_score = 0
                    score_count = 0
                    
                    for exercise in exercises:
                        submission = Submission.find_by_user_and_exercise(student_username, str(exercise['_id']))
                        if submission and submission.get('score') is not None:
                            completed_exercises += 1
                            total_score += submission['score']
                            score_count += 1
                    
                    avg_score = round(total_score / score_count, 1) if score_count > 0 else 0
                    progress_rate = round((completed_exercises / len(exercises)) * 100, 1) if exercises else 0
                    
                    student_stats.append({
                        'username': student_username,
                        'full_name': student.get('full_name', ''),
                        'completed_exercises': completed_exercises,
                        'total_exercises': len(exercises),
                        'progress_rate': progress_rate,
                        'avg_score': avg_score
                    })
            
            return {
                'course': {
                    'id': str(course['_id']),
                    'name': course.get('name'),
                    'total_students': len(students),
                    'total_exercises': len(exercises)
                },
                'exercise_stats': exercise_stats,
                'student_stats': student_stats
            }
        except Exception as e:
            current_app.logger.error(f"Error getting course progress stats: {e}")
            return None
    
    @staticmethod
    def get_recent_activity_stats():
        """
        Lấy thống kê hoạt động gần đây
        """
        username = session.get('username')
        if not username:
            return None
            
        try:
            # Hoạt động 7 ngày gần nhất
            last_7_days = datetime.utcnow() - timedelta(days=7)
            
            # Bài nộp gần đây
            recent_submissions = []
            exercises = list(Exercise.find_by_creator(username))
            for exercise in exercises:
                submissions = list(Submission.find_by_exercise(str(exercise['_id'])))
                for submission in submissions:
                    if submission.get('submitted_at') and submission['submitted_at'] >= last_7_days:
                        student = User.find_by_username(submission['user_id'])
                        recent_submissions.append({
                            'exercise_title': exercise.get('title'),
                            'student_name': student.get('full_name', submission['user_id']) if student else submission['user_id'],
                            'score': submission.get('score'),
                            'submitted_at': submission['submitted_at']
                        })
            
            # Sắp xếp theo thời gian mới nhất
            recent_submissions.sort(key=lambda x: x['submitted_at'], reverse=True)
            
            # Thống kê theo ngày
            daily_stats = {}
            for i in range(7):
                date = datetime.utcnow() - timedelta(days=i)
                date_str = date.strftime('%Y-%m-%d')
                daily_stats[date_str] = {
                    'submissions': 0,
                    'avg_score': 0
                }
            
            for submission in recent_submissions:
                date_str = submission['submitted_at'].strftime('%Y-%m-%d')
                if date_str in daily_stats:
                    daily_stats[date_str]['submissions'] += 1
                    if submission['score'] is not None:
                        daily_stats[date_str]['avg_score'] += submission['score']
            
            # Tính trung bình điểm
            for date_str in daily_stats:
                if daily_stats[date_str]['submissions'] > 0:
                    daily_stats[date_str]['avg_score'] = round(
                        daily_stats[date_str]['avg_score'] / daily_stats[date_str]['submissions'], 1
                    )
            
            return {
                'recent_submissions': recent_submissions[:10],  # Top 10 bài nộp gần nhất
                'daily_stats': daily_stats
            }
        except Exception as e:
            current_app.logger.error(f"Error getting recent activity stats: {e}")
            return None
    
    @staticmethod
    def export_statistics_report(format='excel'):
        """
        Xuất báo cáo thống kê
        """
        username = session.get('username')
        if not username:
            return None
            
        try:
            # Lấy dữ liệu thống kê
            dashboard_stats = StatisticsController.get_lecture_dashboard_stats()
            recent_activity = StatisticsController.get_recent_activity_stats()
            
            # Lấy danh sách khóa học
            courses = list(Course.find_by_creator(username))
            course_stats = []
            
            for course in courses:
                course_progress = StatisticsController.get_course_progress_stats(str(course['_id']))
                if course_progress:
                    course_stats.append(course_progress)
            
            if format == 'excel':
                return StatisticsController._export_excel_report(dashboard_stats, course_stats, recent_activity)
            elif format == 'pdf':
                return StatisticsController._export_pdf_report(dashboard_stats, course_stats, recent_activity)
            else:
                return None
                
        except Exception as e:
            current_app.logger.error(f"Error exporting statistics report: {e}")
            return None
    
    @staticmethod
    def _export_excel_report(dashboard_stats, course_stats, recent_activity):
        """
        Xuất báo cáo Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            
            # Sheet 1: Tổng quan
            ws1 = wb.active
            ws1.title = "Tổng quan"
            
            # Header
            ws1['A1'] = "BÁO CÁO THỐNG KÊ GIẢNG VIÊN"
            ws1['A1'].font = Font(bold=True, size=16)
            ws1.merge_cells('A1:D1')
            
            # Thống kê tổng quan
            row = 3
            ws1[f'A{row}'] = "Thống kê tổng quan"
            ws1[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            stats_data = [
                ["Khóa học", dashboard_stats['courses']['total'], dashboard_stats['courses']['active']],
                ["Sinh viên", dashboard_stats['students']['total'], ""],
                ["Bài tập", dashboard_stats['exercises']['total'], dashboard_stats['exercises']['active']],
                ["Bài nộp", dashboard_stats['submissions']['total'], dashboard_stats['submissions']['recent_7_days']],
                ["Quiz", dashboard_stats['quizzes']['total'], dashboard_stats['quizzes']['submissions']],
                ["Tài liệu", dashboard_stats['documents']['total'], dashboard_stats['documents']['downloads']],
                ["Điểm trung bình", dashboard_stats['performance']['avg_score'], ""]
            ]
            
            ws1['A4'] = "Chỉ số"
            ws1['B4'] = "Tổng số"
            ws1['C4'] = "Chi tiết"
            for cell in ['A4', 'B4', 'C4']:
                ws1[cell].font = Font(bold=True)
            
            for i, (name, total, detail) in enumerate(stats_data):
                row = 5 + i
                ws1[f'A{row}'] = name
                ws1[f'B{row}'] = total
                ws1[f'C{row}'] = detail
            
            # Sheet 2: Thống kê khóa học
            ws2 = wb.create_sheet("Thống kê khóa học")
            ws2['A1'] = "THỐNG KÊ KHÓA HỌC"
            ws2['A1'].font = Font(bold=True, size=16)
            
            row = 3
            ws2['A3'] = "Tên khóa học"
            ws2['B3'] = "Số sinh viên"
            ws2['C3'] = "Số bài tập"
            ws2['D3'] = "Tỷ lệ hoàn thành"
            for cell in ['A3', 'B3', 'C3', 'D3']:
                ws2[cell].font = Font(bold=True)
            
            for course_stat in course_stats:
                row += 1
                ws2[f'A{row}'] = course_stat['course']['name']
                ws2[f'B{row}'] = course_stat['course']['total_students']
                ws2[f'C{row}'] = course_stat['course']['total_exercises']
                
                # Tính tỷ lệ hoàn thành trung bình
                if course_stat['exercise_stats']:
                    avg_completion = sum(ex['completion_rate'] for ex in course_stat['exercise_stats']) / len(course_stat['exercise_stats'])
                    ws2[f'D{row}'] = f"{avg_completion:.1f}%"
            
            # Sheet 3: Hoạt động gần đây
            ws3 = wb.create_sheet("Hoạt động gần đây")
            ws3['A1'] = "HOẠT ĐỘNG 7 NGÀY GẦN ĐÂY"
            ws3['A1'].font = Font(bold=True, size=16)
            
            row = 3
            ws3['A3'] = "Ngày"
            ws3['B3'] = "Số bài nộp"
            ws3['C3'] = "Điểm trung bình"
            for cell in ['A3', 'B3', 'C3']:
                ws3[cell].font = Font(bold=True)
            
            for date_str, stats in recent_activity['daily_stats'].items():
                row += 1
                ws3[f'A{row}'] = date_str
                ws3[f'B{row}'] = stats['submissions']
                ws3[f'C{row}'] = stats['avg_score']
            
            # Lưu file
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
            
        except Exception as e:
            current_app.logger.error(f"Error creating Excel report: {e}")
            return None
    
    @staticmethod
    def _export_pdf_report(dashboard_stats, course_stats, recent_activity):
        """
        Xuất báo cáo PDF (placeholder - có thể implement sau)
        """
        # TODO: Implement PDF export
        return None 